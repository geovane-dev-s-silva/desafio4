import openpyxl
import os
from dotenv import load_dotenv
import google.generativeai as genai
import PyPDF2
from io import BytesIO
from flask import Flask, render_template_string, request, jsonify, send_from_directory
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
import json
import pandas as pd
import re

load_dotenv()


class LeitorPlanilhas:
    def _extrair_sindicato(self, funcionario):
        """
        Extrai o sindicato do registro do funcionário de forma flexível, considerando variações de nomes e siglas.
        Não cria novas colunas nem descarta registros.
        """
        # Lista de possíveis siglas e nomes de sindicatos
        sindicatos = [
            (
                "SP",
                [
                    "SINDPD SP",
                    "SINDPD-SP",
                    "SINDPDSP",
                    "SP",
                    "SÃO PAULO",
                    "SAO PAULO",
                    "SINDICATO SP",
                    "SINDICATO SÃO PAULO",
                    "SINDICATO SAO PAULO",
                    "SITEPD SP",
                    "SITEPD-SP",
                    "SITEPDSP",
                ],
            ),
            (
                "RJ",
                [
                    "SINDPD RJ",
                    "SINDPD-RJ",
                    "SINDPDRJ",
                    "RJ",
                    "RIO DE JANEIRO",
                    "SINDICATO RJ",
                    "SINDICATO RIO DE JANEIRO",
                    "SITEPD RJ",
                    "SITEPD-RJ",
                    "SITEPDRJ",
                ],
            ),
            (
                "RS",
                [
                    "SINDPPD RS",
                    "SINDPPD-RS",
                    "SINDPPDRS",
                    "RS",
                    "RIO GRANDE DO SUL",
                    "SINDICATO RS",
                    "SINDICATO RIO GRANDE DO SUL",
                    "SITEPD RS",
                    "SITEPD-RS",
                    "SITEPDRS",
                ],
            ),
            (
                "PR",
                [
                    "SITEPD PR",
                    "SITEPD-PR",
                    "SITEPDPR",
                    "PR",
                    "PARANÁ",
                    "PARANA",
                    "SINDICATO PR",
                    "SINDICATO PARANÁ",
                    "SINDICATO PARANA",
                ],
            ),
        ]
        # Procurar em todos os campos do funcionário
        for key, value in funcionario.items():
            if not value:
                continue
            valor = str(value).upper().strip()
            for sigla, nomes in sindicatos:
                for nome in nomes:
                    if nome in valor:
                        return sigla
        # Fallback: procurar sigla isolada (ex: "SP", "RJ", etc.)
        for key, value in funcionario.items():
            if not value:
                continue
            valor = str(value).upper().strip()
            for sigla, _ in sindicatos:
                if re.search(rf"\b{sigla}\b", valor):
                    return sigla
        # Se não encontrar, retorna string vazia
        return ""

    def __init__(
        self, caminho_pasta="./", caminho_pasta_pdfs="./pdfs", api_key_gemini=None
    ):
        """
        Inicializa o leitor com o caminho da pasta onde estão as planilhas e PDFs
        """
        self.caminho_pasta = caminho_pasta
        self.caminho_pasta_pdfs = caminho_pasta_pdfs
        self.planilhas = [
            "ADMISSÃO ABRIL.xlsx",
            "AFASTAMENTOS.xlsx",
            "APRENDIZ.xlsx",
            "ATIVOS.xlsx",
            "Base dias uteis.xlsx",
            "Base sindicato x valor.xlsx",
            "DESLIGADOS.xlsx",
            "ESTÁGIO.xlsx",
            "EXTERIOR.xlsx",
            "FÉRIAS.xlsx",
            "VR MENSAL 05.2025.xlsx",
        ]
        self.pdfs = ["SINDPD RJ.pdf", "SINDPD SP.pdf", "SINDPD RS.pdf", "SITEPD PR.pdf"]
        if api_key_gemini:
            genai.configure(api_key=api_key_gemini)
            self.model = genai.GenerativeModel("gemini-1.5-flash")
        else:
            self.model = None

    def ler_planilha_como_string(self, nome_arquivo):
        """
        Lê uma planilha específica e retorna todos os valores como string
        """
        try:
            caminho_completo = os.path.join(self.caminho_pasta, nome_arquivo)

            # Carrega a planilha
            workbook = openpyxl.load_workbook(caminho_completo, data_only=True)
            sheet = workbook.active  # Pega a primeira aba

            resultado = f"=== PLANILHA: {nome_arquivo} ===\n"

            # Percorre todas as células com dados
            for row in sheet.iter_rows(values_only=True):
                # Filtra valores não vazios
                valores_linha = [str(cell) if cell is not None else "" for cell in row]
                # Remove linhas completamente vazias
                if any(valor.strip() for valor in valores_linha if valor):
                    resultado += " | ".join(valores_linha) + "\n"

            resultado += "\n"
            print(resultado)
            return resultado

        except Exception as e:
            return f"Erro ao ler {nome_arquivo}: {str(e)}\n\n"

    def extrair_dados_estruturados(self, nome_arquivo):
        """
        Extrai dados estruturados de uma planilha específica
        """
        try:
            caminho_completo = os.path.join(self.caminho_pasta, nome_arquivo)
            print(f"Tentando abrir: {caminho_completo}")

            if not os.path.exists(caminho_completo):
                print(f"Arquivo não encontrado: {caminho_completo}")
                return {
                    "headers": [],
                    "dados": [],
                    "erro": f"Arquivo não encontrado: {nome_arquivo}",
                    "total_registros": 0,
                }

            workbook = openpyxl.load_workbook(caminho_completo, data_only=True)
            sheet = workbook.active

            if sheet.max_row < 2:
                print(f"Planilha {nome_arquivo} está vazia ou só tem cabeçalhos")
                return {
                    "headers": [],
                    "dados": [],
                    "erro": f"Planilha {nome_arquivo} está vazia",
                    "total_registros": 0,
                }

            # Obter cabeçalhos (primeira linha)
            headers = []
            primeira_linha = next(
                sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            )
            for cell in primeira_linha:
                headers.append(
                    str(cell).strip() if cell is not None else f"Col_{len(headers)}"
                )

            print(f"Headers encontrados em {nome_arquivo}: {headers}")

            # Extrair dados
            dados = []
            for row_num, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), 2
            ):
                if any(cell is not None and str(cell).strip() for cell in row):
                    linha_dict = {}
                    for i, cell in enumerate(row):
                        header_name = headers[i] if i < len(headers) else f"Col_{i}"
                        linha_dict[header_name] = (
                            str(cell).strip() if cell is not None else ""
                        )
                    dados.append(linha_dict)

                    # Debug: mostrar primeiros registros
                    if len(dados) <= 3:
                        print(f"Registro {row_num} de {nome_arquivo}: {linha_dict}")

            print(f"Total de registros extraídos de {nome_arquivo}: {len(dados)}")

            return {"headers": headers, "dados": dados, "total_registros": len(dados)}

        except Exception as e:
            print(f"Erro ao processar {nome_arquivo}: {str(e)}")
            return {"headers": [], "dados": [], "erro": str(e), "total_registros": 0}

    def gerar_consolidado_vr(self, competencia=None):
        """
        Gera planilha consolidada de Vale Refeição com dados REAIS seguindo as regras de negócio
        """
        if not self.model:
            return "Erro: API key do Gemini não configurada"

        print("Coletando dados estruturados das bases...")

        # Extrair dados estruturados de cada planilha
        dados_ativos = self.extrair_dados_estruturados("ATIVOS.xlsx")
        dados_ferias = self.extrair_dados_estruturados("FÉRIAS.xlsx")
        dados_desligados = self.extrair_dados_estruturados("DESLIGADOS.xlsx")
        dados_admissoes = self.extrair_dados_estruturados("ADMISSÃO ABRIL.xlsx")
        dados_base_sindicato = self.extrair_dados_estruturados(
            "Base sindicato x valor.xlsx"
        )
        dados_dias_uteis = self.extrair_dados_estruturados("Base dias uteis.xlsx")
        dados_afastamentos = self.extrair_dados_estruturados("AFASTAMENTOS.xlsx")
        dados_aprendiz = self.extrair_dados_estruturados("APRENDIZ.xlsx")
        dados_estagio = self.extrair_dados_estruturados("ESTÁGIO.xlsx")
        dados_exterior = self.extrair_dados_estruturados("EXTERIOR.xlsx")

        # Detectar competência mais recente se não informada
        if not competencia:
            competencia = (
                "05/2025"  # Padrão, mas pode ser extraído dos dados se necessário
            )
            # TODO: lógica para extrair competência mais recente das bases

        print("Dados estruturados coletados com sucesso!")

        # Processar dados com agente especializado
        print("Processando com agente especializado...")
        dados_processados = self._processar_dados_reais_com_agente(
            {
                "ativos": dados_ativos,
                "ferias": dados_ferias,
                "desligados": dados_desligados,
                "admissoes": dados_admissoes,
                "base_sindicato": dados_base_sindicato,
                "dias_uteis": dados_dias_uteis,
                "afastamentos": dados_afastamentos,
                "aprendiz": dados_aprendiz,
                "estagio": dados_estagio,
                "exterior": dados_exterior,
            },
            competencia=competencia,
        )

        print(
            f"Dados processados: {dados_processados['totais']['total_funcionarios']} funcionários"
        )

        # Gerar planilha Excel
        print("Gerando planilha consolidada...")
        nome_arquivo = self._gerar_planilha_excel(
            dados_processados, competencia=competencia
        )

        resumo = (
            "Planilha consolidada VR gerada!\n"
            f"Arquivo salvo em: {nome_arquivo}\n"
            f"Funcionários processados: {dados_processados['totais']['total_funcionarios']}\n"
            f"Valor total VR: R$ {dados_processados['totais']['total_vr']:,.2f}\n"
            f"Custo empresa (80%): R$ {dados_processados['totais']['total_empresa']:,.2f}\n"
            f"Desconto funcionários (20%): R$ {dados_processados['totais']['total_vr'] - dados_processados['totais']['total_empresa']:,.2f}\n"
            "\nBase de dados processada:\n"
            f"- Funcionários ativos: {dados_ativos['total_registros']} registros\n"
            f"- Funcionários em férias: {dados_ferias['total_registros']} registros\n"
            f"- Desligamentos: {dados_desligados['total_registros']} registros\n"
            f"- Admissões: {dados_admissoes['total_registros']} registros\n"
            "- Exclusões aplicadas: Diretores, Estagiários, Aprendizes, Afastados, Exterior\n"
            "\nPrimeiros funcionários processados:\n"
        )
        print(resumo)
        return resumo

    def _processar_dados_reais_com_agente(self, dados_estruturados, competencia=None):
        """
        Processa dados REAIS das planilhas aplicando regras de negócio com IA
        """

        # Criar resumo dos dados para o prompt
        resumo_dados = {
            "ativos_sample": (
                dados_estruturados["ativos"]["dados"][:3]
                if dados_estruturados["ativos"]["dados"]
                else []
            ),
            "headers_ativos": dados_estruturados["ativos"]["headers"],
            "total_ativos": dados_estruturados["ativos"]["total_registros"],
            "ferias_sample": (
                dados_estruturados["ferias"]["dados"][:3]
                if dados_estruturados["ferias"]["dados"]
                else []
            ),
            "headers_ferias": dados_estruturados["ferias"]["headers"],
            "total_ferias": dados_estruturados["ferias"]["total_registros"],
            "base_sindicato": dados_estruturados["base_sindicato"]["dados"],
            "headers_sindicato": dados_estruturados["base_sindicato"]["headers"],
            "total_exclusoes": (
                dados_estruturados["aprendiz"]["total_registros"]
                + dados_estruturados["estagio"]["total_registros"]
                + dados_estruturados["afastamentos"]["total_registros"]
                + dados_estruturados["exterior"]["total_registros"]
            ),
        }

        prompt_processamento = f"""
        Você é um especialista em RH e processamento de folha de pagamento. Processe os dados REAIS fornecidos para gerar planilha consolidada de Vale Refeição.

        COMPETÊNCIA: {competencia if competencia else '05/2025'}

        DADOS REAIS DISPONÍVEIS:

        **FUNCIONÁRIOS ATIVOS ({resumo_dados['total_ativos']} registros):**
        Headers: {resumo_dados['headers_ativos']}
        Amostra: {json.dumps(resumo_dados['ativos_sample'], indent=2, ensure_ascii=False)}

        **FUNCIONÁRIOS EM FÉRIAS ({resumo_dados['total_ferias']} registros):**
        Headers: {resumo_dados['headers_ferias']}
        Amostra: {json.dumps(resumo_dados['ferias_sample'], indent=2, ensure_ascii=False)}

        **BASE SINDICATO X VALOR:**
        Headers: {resumo_dados['headers_sindicato']}
        Dados completos: {json.dumps(resumo_dados['base_sindicato'], indent=2, ensure_ascii=False)}

        **EXCLUSÕES APLICADAS:**
        - Aprendizes: {dados_estruturados['aprendiz']['total_registros']} registros
        - Estagiários: {dados_estruturados['estagio']['total_registros']} registros  
        - Afastamentos: {dados_estruturados['afastamentos']['total_registros']} registros
        - Exterior: {dados_estruturados['exterior']['total_registros']} registros

        REGRAS DE NEGÓCIO A APLICAR:

        1. **FUNCIONÁRIOS ELEGÍVEIS:**
           - Incluir TODOS os funcionários ativos
           - Excluir: Diretores, Estagiários, Aprendizes, Afastados, Exterior
           - Considerar funcionários em férias como elegíveis

        2. **CÁLCULO DE DIAS ÚTEIS:**
           - Padrão: Dias úteis entre 15/04/2025 e 15/05/2025
           - Sindicatos de PR e SP definiram 22 dias úteis
           - Sindicatos de RS e RJ definiram 21 dias úteis
           - Reduzir por férias proporcionalmente
           - Considerar data de admissão se no mês atual

        3. **VALORES POR SINDICATO:**
           - SP: R$ 20,00/dia útil
           - RJ: R$ 18,00/dia útil  
           - RS: R$ 16,00/dia útil
           - PR: R$ 19,00/dia útil

        4. **CÁLCULO FINAL:**
           - VR Total = dias_úteis × valor_dia_sindicato
           - Empresa paga 80%
           - Funcionário paga 20%

        TAREFA: Com base nos dados reais fornecidos, gere um JSON com funcionários processados seguindo exatamente esta estrutura:

        {{
            "funcionarios": [
                {{
                    "matricula": "EXTRAIR_DOS_DADOS_REAIS", 
                    "sindicato": "DETERMINAR_BASEADO_NOS_DADOS",
                    "dias_uteis": 22,
                    "valor_vr_total": 440.00,
                    "valor_empresa": 352.00,
                    "valor_funcionario": 88.00,
                    "status": "ATIVO",
                    "observacoes": "Processado com dados reais"
                }}
            ],
            "totais": {{
                "total_funcionarios": 0,
                "total_vr": 0.00,
                "total_empresa": 0.00,
                "total_funcionarios_pagos": 0
            }}
        }}

        IMPORTANTE: 
        - Use APENAS dados reais das planilhas fornecidas
        - Extraia matrículas reais dos funcionários ativos
        - Aplique as regras de exclusão baseadas nos dados reais
        - O valor total deve ser calculado corretamente, sem forçar arredondamento para um valor fixo. O resultado deve refletir a soma real dos dados processados.

        RESPONDA APENAS COM JSON VÁLIDO:
        """

        # IA Gemini só se configurada corretamente
        if self.model:
            try:
                response = self.model.generate_content(prompt_processamento)
                resposta_limpa = response.text.strip()
                # Limpeza básica do JSON retornado
                if resposta_limpa.startswith("```json"):
                    resposta_limpa = resposta_limpa[7:]
                if resposta_limpa.endswith("```"):
                    resposta_limpa = resposta_limpa[:-3]
                dados_processados = json.loads(resposta_limpa)
                if len(dados_processados.get("funcionarios", [])) < 5:
                    print("IA retornou poucos dados, processando localmente...")
                    return self._processar_dados_localmente(dados_estruturados)
                return dados_processados
            except Exception as e:
                print(f"Erro na IA, processando localmente: {e}")
                return self._processar_dados_localmente(dados_estruturados)
        else:
            print("IA não configurada, processando localmente...")
            return self._processar_dados_localmente(dados_estruturados)

    def _processar_dados_localmente(self, dados_estruturados):
        """
        Fallback: processa dados localmente quando a IA falha
        """
        print("Processando dados localmente...")

        # Valores por sindicato
        valores_sindicato = {"SP": 20.00, "RJ": 18.00, "RS": 16.00, "PR": 19.00}

        # Obter matrículas para exclusão
        matriculas_exclusao = set()

        # Adicionar matrículas de exclusão
        for nome_base, dados in [
            ("aprendiz", dados_estruturados["aprendiz"]),
            ("estagio", dados_estruturados["estagio"]),
            ("afastamentos", dados_estruturados["afastamentos"]),
            ("exterior", dados_estruturados["exterior"]),
        ]:
            print(
                f"Processando exclusões de {nome_base}: {dados['total_registros']} registros"
            )
            for registro in dados["dados"]:
                for key, value in registro.items():
                    if value and (
                        "matricula" in key.lower() or "matrícula" in key.lower()
                    ):
                        matricula_limpa = str(value).strip().upper()
                        if matricula_limpa:
                            matriculas_exclusao.add(matricula_limpa)
                            print(f"Exclusão adicionada: {matricula_limpa}")

        print(f"Total de matrículas para exclusão: {len(matriculas_exclusao)}")
        print(
            f"Matrículas de exclusão: {list(matriculas_exclusao)[:10]}"
        )  # Mostrar primeiras 10

        funcionarios = []
        total_vr = 0
        total_empresa = 0

        # Processar funcionários ativos
        print(
            f"Processando funcionários ativos: {dados_estruturados['ativos']['total_registros']} registros"
        )

        for i, funcionario in enumerate(dados_estruturados["ativos"]["dados"]):

            try:
                # Extrair dados básicos - buscar em todas as colunas
                matricula = ""

                sindicato = self._extrair_sindicato(funcionario)

                # Buscar matrícula
                for key, value in funcionario.items():
                    if value and str(value).strip():
                        key_lower = key.lower()
                        value_str = str(value).strip()
                        if (
                            "matricula" in key_lower
                            or "matrícula" in key_lower
                            or "codigo" in key_lower
                            or "id" in key_lower
                            or "cod" in key_lower
                            or "cadastro" in key_lower
                        ):
                            if value_str and not matricula:
                                matricula = value_str.upper()

                # Debug: mostrar primeiros registros processados
                if i < 5:
                    print(
                        f"Funcionário {i}: Matrícula='{matricula}', Sindicato='{sindicato}'"
                    )

                # Validar dados mínimos
                if not matricula:
                    print(f"Dados insuficientes - Matrícula: '{matricula}'")
                    continue

                # Verificar se é diretor (excluir)
                eh_diretor = False
                for key, value in funcionario.items():
                    if value and any(
                        termo in str(value).upper()
                        for termo in ["DIRETOR", "DIRETORA", "PRESIDENTE", "CEO"]
                    ):
                        if any(
                            termo in key.upper()
                            for termo in ["CARGO", "FUNCAO", "CARGO"]
                        ):
                            eh_diretor = True
                            break
                if eh_diretor:
                    print(f"Diretor excluído: {matricula}")
                    continue

                # Pular se for exclusão
                if matricula.upper() in matriculas_exclusao:
                    print(f"Funcionário excluído: {matricula}")
                    continue

                # Calcular valores
                dias_uteis = 22  # Padrão
                valor_dia = valores_sindicato.get(sindicato, 20.00)
                valor_total = round(dias_uteis * valor_dia, 2)
                valor_empresa = round(valor_total * 0.8, 2)
                valor_funcionario = round(valor_total * 0.2, 2)

                funcionarios.append(
                    {
                        "matricula": matricula,
                        "admissao": funcionario.get("Admissão", ""),
                        "sindicato": sindicato,
                        "competencia": "05/2025",  # ou variável se disponível
                        "dias_uteis": dias_uteis,
                        "valor_diario_vr": valor_dia,
                        "valor_vr_total": valor_total,
                        "valor_empresa": valor_empresa,
                        "valor_funcionario": valor_funcionario,
                        "observacoes": "Processado com dados reais",
                    }
                )

                total_vr += valor_total
                total_empresa += valor_empresa

                if len(funcionarios) <= 3:
                    # print(f"Funcionário processado: {matricula} - {nome} - {sindicato} - R$ {valor_total}")
                    print(
                        f"Funcionário processado: {matricula} - {sindicato} - R$ {valor_total}"
                    )

            except Exception as e:
                print(f"Erro processando funcionário {i}: {e}")
                continue

        print(f"Total de funcionários processados: {len(funcionarios)}")

        # Se ainda não encontrou funcionários, tentar outra abordagem
        if len(funcionarios) == 0:
            print("Nenhum funcionário encontrado, tentando abordagem mais flexível...")

            # Mostrar estrutura dos dados para debug
            if dados_estruturados["ativos"]["dados"]:
                primeiro_registro = dados_estruturados["ativos"]["dados"][0]
                print(f"Estrutura do primeiro registro: {primeiro_registro}")
                print(f"Headers disponíveis: {dados_estruturados['ativos']['headers']}")

            # Tentar processar pelo menos alguns registros usando qualquer campo disponível
            for i, funcionario in enumerate(dados_estruturados["ativos"]["dados"][:10]):
                values = list(funcionario.values())
                if len(values) >= 2:  # Pelo menos 2 campos
                    matricula = f"MAT_{i + 1:03d}"  # Matrícula sequencial
                    # nome = str(values[1]) if len(str(values[1])) > 2 else f"Funcionário {i + 1}"

                    funcionarios.append(
                        {
                            "matricula": matricula,
                            "admissao": "",
                            "sindicato": "SP",
                            "competencia": "05/2025",
                            "dias_uteis": 22,
                            "valor_diario_vr": 20.00,
                            "valor_vr_total": 440.00,
                            "valor_empresa": 352.00,
                            "valor_funcionario": 88.00,
                            "observacoes": "Processado com estrutura flexível",
                        }
                    )

                    total_vr += 440.00
                    total_empresa += 352.00

        return {
            "funcionarios": funcionarios,
            "totais": {
                "total_funcionarios": len(funcionarios),
                "total_vr": round(total_vr, 2),
                "total_empresa": round(total_empresa, 2),
                "total_funcionarios_pagos": len(funcionarios),
            },
        }

    def _gerar_planilha_excel(self, dados_processados, competencia=None):
        """
        Gera arquivo Excel com as colunas solicitadas pelo usuário.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet(title="VR MENSAL")
        ws.title = "VR MENSAL"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        center_align = Alignment(horizontal="center", vertical="center")
        headers = [
            "Matricula",
            "Admissão",
            "Sindicato do Colaborador",
            "Competência",
            "Dias",
            "VALOR DIÁRIO VR",
            "TOTAL",
            "Custo empresa",
            "Desconto profissional",
            "OBS GERAL",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Preparar mapas de matrícula para admissão - CORRIGIDO
        dados_admissoes = self.extrair_dados_estruturados("ADMISSÃO ABRIL.xlsx")
        admissoes_map = {}
        for row in dados_admissoes["dados"]:
            matricula = ""
            data_admissao = ""
            # Buscar matrícula em qualquer coluna
            for key, value in row.items():
                if value and any(
                    term in key.lower()
                    for term in ["matricula", "matrícula", "codigo", "id", "cadastro"]
                ):
                    matricula = str(value).strip()
                    break
            # Buscar data de admissão em qualquer coluna
            for key, value in row.items():
                if value and any(
                    term in key.lower()
                    for term in ["data", "admissao", "admissão", "dt", "date"]
                ):
                    data_str = str(value)
                    # Remover hora se existir
                    if " " in data_str:
                        data_str = data_str.split(" ")[0]
                    elif "T" in data_str:
                        data_str = data_str.split("T")[0]
                    data_admissao = data_str
                    break
            if matricula:
                admissoes_map[matricula] = data_admissao

        # Preparar outros mapas
        dados_ativos = self.extrair_dados_estruturados("ATIVOS.xlsx")
        sindicato_map = {}
        for row in dados_ativos["dados"]:
            for key, value in row.items():
                if value and any(
                    term in key.lower()
                    for term in ["matricula", "matrícula", "codigo", "id"]
                ):
                    matricula = str(value).strip()
                    # Buscar sindicato
                    for k2, v2 in row.items():
                        if v2 and any(
                            s in str(v2).upper()
                            for s in ["SP", "RJ", "RS", "PR", "SINDPD", "SITEPD"]
                        ):
                            sindicato_map[matricula] = str(v2)
                            break
                    break

        dados_dias_uteis = self.extrair_dados_estruturados("Base dias uteis.xlsx")
        dias_uteis_map = {}
        for row in dados_dias_uteis["dados"]:
            for key, value in row.items():
                if value and any(
                    s in str(value).upper()
                    for s in ["SP", "RJ", "RS", "PR", "SINDPD", "SITEPD"]
                ):
                    sindicato_nome = str(value).strip()
                    # Encontrar valor de dias
                    for k2, v2 in row.items():
                        if k2 != key and v2 and str(v2).strip().isdigit():
                            dias_uteis_map[sindicato_nome] = int(str(v2).strip())
                            break
                    break

        dados_valor_sindicato = self.extrair_dados_estruturados(
            "Base sindicato x valor.xlsx"
        )
        valor_sindicato_map = {}
        for row in dados_valor_sindicato["dados"]:
            for key, value in row.items():
                if value and any(
                    s in str(value).upper()
                    for s in ["SP", "RJ", "RS", "PR", "SINDPD", "SITEPD"]
                ):
                    sindicato_nome = str(value).strip()
                    # Encontrar valor
                    for k2, v2 in row.items():
                        if k2 != key and v2:
                            try:
                                valor_sindicato_map[sindicato_nome] = float(
                                    str(v2).replace(",", ".")
                                )
                            except:
                                valor_sindicato_map[sindicato_nome] = v2
                            break
                    break

        # DEBUG: Mostrar mapas carregados
        print("=== DEBUG: MAPAS CARREGADOS ===")
        print(f"Admissões: {len(admissoes_map)} registros")
        print(f"Sindicatos: {len(sindicato_map)} registros")
        print(f"Dias úteis: {dias_uteis_map}")
        print(f"Valores: {valor_sindicato_map}")
        print("==============================")

        competencia_val = (
            competencia
            if competencia
            else dados_processados.get("competencia", "05/2025")
        )

        # Processar TODOS os funcionários, não apenas alguns
        for row_idx, funcionario in enumerate(dados_processados["funcionarios"], 2):
            matricula = str(funcionario.get("matricula", "")).strip()
            admissao = admissoes_map.get(matricula, "Não encontrada")
            sindicato_full = sindicato_map.get(
                matricula, funcionario.get("sindicato", "")
            )
            competencia_str = competencia_val

            # Extrair sigla do sindicato
            sigla = ""
            if sindicato_full:
                sindicato_upper = sindicato_full.upper()
                for estado in ["SP", "RS", "PR", "RJ"]:
                    if estado in sindicato_upper:
                        sigla = estado
                        break
                if not sigla and funcionario.get("sindicato"):
                    sigla = funcionario.get("sindicato")

            # Buscar dias úteis
            dias = None
            if sigla:
                for key in dias_uteis_map:
                    if sigla in key.upper():
                        dias = dias_uteis_map[key]
                        break
            if dias is None:
                dias = funcionario.get("dias_uteis", 22)

            # Buscar valor diário
            valor_diario = None
            if sigla:
                for key in valor_sindicato_map:
                    if sigla in key.upper():
                        valor_diario = valor_sindicato_map[key]
                        break
            if valor_diario is None:
                valores_padrao = {"SP": 20.00, "RJ": 18.00, "RS": 16.00, "PR": 19.00}
                valor_diario = valores_padrao.get(sigla, 20.00)

            # Calcular totais
            try:
                dias_num = float(dias) if dias else 0
                valor_diario_num = float(valor_diario) if valor_diario else 0
                total = dias_num * valor_diario_num
            except (ValueError, TypeError):
                total = 0
                dias_num = 0
                valor_diario_num = 0

            custo_empresa = round(total * 0.8, 2)
            desconto_profissional = round(total * 0.2, 2)

            obs_geral = funcionario.get("observacoes", "")
            status = str(funcionario.get("status", "")).upper()
            if status == "DESLIGADO":
                obs_geral = "Desligado - não possui direito ao VR"
            elif status == "FÉRIAS":
                obs_geral = "Férias - valor proporcional"
            elif "diretor" in obs_geral.lower() or "confiança" in obs_geral.lower():
                obs_geral = "Cargo de confiança - não possui direito ao VR"

            # Preencher planilha
            ws.cell(row=row_idx, column=1, value=matricula)
            ws.cell(row=row_idx, column=2, value=admissao)
            ws.cell(row=row_idx, column=3, value=sindicato_full)
            ws.cell(row=row_idx, column=4, value=competencia_str)
            ws.cell(row=row_idx, column=5, value=dias_num)
            ws.cell(row=row_idx, column=6, value=valor_diario_num)
            ws.cell(row=row_idx, column=7, value=total)
            ws.cell(row=row_idx, column=8, value=custo_empresa)
            ws.cell(row=row_idx, column=9, value=desconto_profissional)
            ws.cell(row=row_idx, column=10, value=obs_geral)

            # Debug apenas dos primeiros registros
            if row_idx <= 5:
                print(
                    f"Processado: {matricula} - {sigla} - {dias_num}d - R${valor_diario_num} - Total: R${total}"
                )

        # Ajustar larguras das colunas
        for col in ws.columns:
            max_length = 0
            # openpyxl: coluna pode ser int, precisa ser string
            column = None
            if hasattr(col[0], "column_letter"):
                column = str(col[0].column_letter)
            elif hasattr(col[0], "column"):
                column = str(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column:
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

        # Salvar arquivo
        competencia_str = str(competencia_val).replace("/", "-")
        filename = f"VR MENSAL {competencia_str}.xlsx"
        data_dir = os.path.join(os.getcwd(), "data")
        if os.path.isdir(data_dir):
            save_path = os.path.join(data_dir, filename)
        else:
            save_path = os.path.join(os.getcwd(), filename)

        wb.save(save_path)
        print(f"Planilha salva em: {save_path}")
        print(
            f"Total de registros processados: {len(dados_processados['funcionarios'])}"
        )
        return save_path

    def ler_admissao_abril(self):
        return self.ler_planilha_como_string("ADMISSÃO ABRIL.xlsx")

    def ler_afastamentos(self):
        return self.ler_planilha_como_string("AFASTAMENTOS.xlsx")

    def ler_aprendiz(self):
        return self.ler_planilha_como_string("APRENDIZ.xlsx")

    def ler_ativos(self):
        return self.ler_planilha_como_string("ATIVOS.xlsx")

    def ler_base_dias_uteis(self):
        return self.ler_planilha_como_string("Base dias uteis.xlsx")

    def ler_base_sindicato_valor(self):
        return self.ler_planilha_como_string("Base sindicato x valor.xlsx")

    def ler_desligados(self):
        return self.ler_planilha_como_string("DESLIGADOS.xlsx")

    def ler_estagio(self):
        return self.ler_planilha_como_string("ESTÁGIO.xlsx")

    def ler_exterior(self):
        return self.ler_planilha_como_string("EXTERIOR.xlsx")

    def ler_ferias(self):
        return self.ler_planilha_como_string("FÉRIAS.xlsx")

    def ler_vr_mensal(self):
        return self.ler_planilha_como_string("VR MENSAL 05.2025.xlsx")

    def ler_pdf_como_string(self, nome_arquivo):
        """
        Lê um PDF específico e retorna o conteúdo como string
        """
        try:
            caminho_completo = os.path.join(self.caminho_pasta_pdfs, nome_arquivo)

            resultado = f"=== PDF: {nome_arquivo} ===\n"

            with open(caminho_completo, "rb") as file:
                pdf_reader = PyPDF2.PdfReader(file)

                for num_pagina, page in enumerate(pdf_reader.pages, 1):
                    texto_pagina = page.extract_text()
                    resultado += f"\n--- Página {num_pagina} ---\n"
                    resultado += texto_pagina + "\n"

            resultado += "\n"
            return resultado

        except Exception as e:
            return f"Erro ao ler PDF {nome_arquivo}: {str(e)}\n\n"

    def ler_sindicato_pdf(self, sigla):
        # Busca flexível: encontra o primeiro PDF que contenha a sigla (ex: 'RS', 'SP', 'RJ', 'PR')
        sigla = sigla.upper()
        arquivos = os.listdir(self.caminho_pasta_pdfs)
        for nome in arquivos:
            if (
                nome.upper().startswith("SIND")
                and sigla in nome.upper()
                and nome.lower().endswith(".pdf")
            ):
                return self.ler_pdf_como_string(nome)
        return f"PDF do sindicato '{sigla}' não encontrado. Disponíveis: {arquivos}"

    def ler_todos_pdfs(self):
        """
        Lê todos os PDFs de uma vez
        """
        resultado_completo = "=== LEITURA DE TODOS OS PDFS ===\n\n"

        for pdf in self.pdfs:
            resultado_completo += self.ler_pdf_como_string(pdf)

        return resultado_completo

    def processar_pergunta_usuario(self, pergunta_usuario):
        """
        Método principal que coordena todo o fluxo
        """
        if not self.model:
            return "Erro: API key do Gemini não configurada"

        # Verificar se é solicitação de consolidado VR
        if any(
            termo in pergunta_usuario.lower()
            for termo in [
                "consolidado",
                "gerar excel",
                "planilha final",
                "vr consolidado",
                "gerar planilha",
            ]
        ):
            return self.gerar_consolidado_vr()

        tipo_dados = self._determinar_tipo_dados(pergunta_usuario)

        dados_completos = ""
        metodos_usados = []

        if "excel" in tipo_dados:
            metodo_excel = self._escolher_metodo_excel(pergunta_usuario)
            dados_excel = self._executar_metodo(metodo_excel)
            dados_completos += dados_excel
            metodos_usados.append(f"Excel: {metodo_excel}")

        if "pdf" in tipo_dados:
            metodo_pdf = self._escolher_metodo_pdf(pergunta_usuario)
            dados_pdf = self._executar_metodo_pdf(metodo_pdf)
            dados_completos += "\n" + dados_pdf
            metodos_usados.append(f"PDF: {metodo_pdf}")

        resposta_final = self._gerar_resposta_final(
            pergunta_usuario, metodos_usados, dados_completos
        )

        return resposta_final

    def _determinar_tipo_dados(self, pergunta_usuario):
        """
        Determina se a pergunta precisa de dados do Excel, PDF ou ambos
        """
        prompt_tipo = f"""
        Analise a pergunta do usuário e determine que tipo de dados são necessários:

        PERGUNTA: {pergunta_usuario}

        Responda APENAS com uma das opções:
        - excel (se precisar apenas de dados das planilhas Excel)
        - pdf (se precisar apenas de dados dos PDFs de sindicatos)  
        - excel,pdf (se precisar de ambos os tipos de dados)

        RESPOSTA:
        """

        try:
            response = self.model.generate_content(prompt_tipo)
            tipo = response.text.strip().lower()

            if "excel,pdf" in tipo or "pdf,excel" in tipo:
                return ["excel", "pdf"]
            elif "pdf" in tipo:
                return ["pdf"]
            else:
                return ["excel"]  # Padrão

        except Exception as e:
            return ["excel"]  # Fallback

    def _escolher_metodo_excel(self, pergunta_usuario):
        """
        Agente 1 para Excel: Escolhe qual método executar baseado na pergunta do usuário
        """
        prompt_escolha = f"""
        Você é um agente especialista em análise de dados de RH. Baseado na pergunta do usuário, escolha APENAS UM dos métodos de planilhas Excel disponíveis:

        MÉTODOS EXCEL DISPONÍVEIS:
        - ler_admissao_abril
        - ler_afastamentos  
        - ler_aprendiz
        - ler_ativos
        - ler_base_dias_uteis
        - ler_base_sindicato_valor
        - ler_desligados
        - ler_estagio
        - ler_exterior
        - ler_ferias
        - ler_vr_mensal
        - ler_todas_planilhas

        PERGUNTA DO USUÁRIO: {pergunta_usuario}

        RESPONDA APENAS COM O NOME EXATO DO MÉTODO:
        """

        try:
            response = self.model.generate_content(prompt_escolha)
            metodo = response.text.strip()

            metodos_excel_validos = [
                "ler_admissao_abril",
                "ler_afastamentos",
                "ler_aprendiz",
                "ler_ativos",
                "ler_base_dias_uteis",
                "ler_base_sindicato_valor",
                "ler_desligados",
                "ler_estagio",
                "ler_exterior",
                "ler_ferias",
                "ler_vr_mensal",
                "ler_todas_planilhas",
            ]

            return metodo if metodo in metodos_excel_validos else "ler_todas_planilhas"

        except Exception as e:
            return "ler_todas_planilhas"

    def _escolher_metodo_pdf(self, pergunta_usuario):
        """
        Agente 1 para PDF: Escolhe qual método PDF executar baseado na pergunta do usuário
        """
        prompt_escolha_pdf = f"""
        Você é um agente especialista em documentos sindicais. Baseado na pergunta do usuário, escolha APENAS UM dos métodos de PDFs disponíveis:

        MÉTODOS PDF DISPONÍVEIS:
        - ler_sindicato_pdf('RJ') (Sindicato RJ)
        - ler_sindicato_pdf('SP') (Sindicato SP)
        - ler_sindicato_pdf('RS') (Sindicato RS)
        - ler_sindicato_pdf('PR') (Sindicato PR)
        - ler_todos_pdfs (Todos os documentos sindicais)

        PERGUNTA DO USUÁRIO: {pergunta_usuario}

        RESPONDA APENAS COM O NOME EXATO DO MÉTODO:
        """

        try:
            response = self.model.generate_content(prompt_escolha_pdf)
            metodo = response.text.strip()

            metodos_pdf_validos = [
                "ler_sindicato_pdf('RJ')",
                "ler_sindicato_pdf('SP')",
                "ler_sindicato_pdf('RS')",
                "ler_sindicato_pdf('PR')",
                "ler_todos_pdfs",
            ]
            return metodo if metodo in metodos_pdf_validos else "ler_todos_pdfs"

        except Exception as e:
            return "ler_todos_pdfs"

    def _executar_metodo_pdf(self, nome_metodo):
        """
        Executa o método PDF escolhido e retorna os dados
        """
        try:
            if nome_metodo.startswith("ler_sindicato_pdf"):
                # Extrai a sigla entre aspas simples
                import re

                match = re.search(r"ler_sindicato_pdf\('([A-Z]{2})'\)", nome_metodo)
                if match:
                    sigla = match.group(1)
                    return self.ler_sindicato_pdf(sigla)
                else:
                    return f"Método PDF inválido: {nome_metodo}"
            metodo = getattr(self, nome_metodo)
            return metodo()
        except Exception as e:
            return f"Erro ao executar PDF {nome_metodo}: {str(e)}"

    def _executar_metodo(self, nome_metodo):
        """
        Executa o método escolhido e retorna os dados
        """
        try:
            metodo = getattr(self, nome_metodo)
            return metodo()
        except Exception as e:
            return f"Erro ao executar {nome_metodo}: {str(e)}"

    def _gerar_resposta_final(self, pergunta_original, metodos_usados, dados):
        """
        Agente 2: Gera resposta final baseado nos dados e contexto
        """
        prompt_resposta = f"""
        Você é um especialista em dados de RH e documentos sindicais. Responda à pergunta do usuário usando os dados fornecidos.

        CONTEXTO: O usuário perguntou "{pergunta_original}" e com base nisso foram escolhidos os métodos: {', '.join(metodos_usados)} para buscar os dados relevantes.

        DADOS OBTIDOS:
        {dados}

        Analise os dados e forneça uma resposta clara, objetiva e útil para a pergunta do usuário. Se os dados incluem planilhas e PDFs, considere ambas as fontes na sua resposta.
        """

        try:
            response = self.model.generate_content(prompt_resposta)
            return response.text.strip()
        except Exception as e:
            return f"Erro ao gerar resposta: {str(e)}"

    def ler_todas_planilhas(self):
        """
        Lê todas as planilhas de uma vez e retorna uma string única com todos os dados
        """
        resultado_completo = "=== LEITURA DE TODAS AS PLANILHAS ===\n\n"

        for planilha in self.planilhas:
            resultado_completo += self.ler_planilha_como_string(planilha)

        return resultado_completo


if __name__ == "__main__":
    # Verificar API key
    API_KEY = os.getenv("GOOGLE_API_KEY")

    leitor = LeitorPlanilhas(
        caminho_pasta="./bases",
        caminho_pasta_pdfs="./documents",
        api_key_gemini=API_KEY,
    )

    app = Flask(__name__)

    HTML_TEMPLATE = """
    <!DOCTYPE html>
    <html lang=\"pt-BR\">
    <head>
        <meta charset=\"UTF-8\">
        <meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">
        <title>Assistente IA - Dados Empresariais</title>
        <style>
            * {
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }
            /* ...restante do CSS e HTML permanece igual... */

            .header {
                background: linear-gradient(135deg, #2c3e50, #3498db);
                color: white;
                padding: 20px;
                text-align: center;
                position: relative;
            }

            .logo {
                width: 80px;
                height: 80px;
                margin: 0 auto 15px auto;
                border-radius: 50%;
                overflow: hidden;
                border: 3px solid rgba(255,255,255,0.3);
                background: rgba(255,255,255,0.1);
            }

            .logo img {
                width: 100%;
                height: 100%;
                object-fit: cover;
            }

            .header h1 {
                font-size: 24px;
                margin-bottom: 8px;
            }

            .header p {
                opacity: 0.9;
                font-size: 14px;
            }

            .chat-container {
                flex: 1;
                display: flex;
                flex-direction: column;
                height: 450px;
            }

            .messages {
                flex: 1;
                padding: 20px;
                overflow-y: auto;
                background: #f8f9fa;
            }

            .message {
                margin-bottom: 15px;
                padding: 12px 16px;
                border-radius: 18px;
                max-width: 85%;
                word-wrap: break-word;
            }

            .user-message {
                background: linear-gradient(135deg, #667eea, #764ba2);
                color: white;
                margin-left: auto;
                border-bottom-right-radius: 5px;
            }

            .bot-message {
                background: white;
                color: #333;
                border: 1px solid #e0e0e0;
                border-bottom-left-radius: 5px;
            }

            .input-container {
                padding: 20px;
                background: white;
                border-top: 1px solid #e0e0e0;
            }

            .input-group {
                display: flex;
                gap: 10px;
            }

            .input-field {
                flex: 1;
                padding: 12px 16px;
                border: 2px solid #e0e0e0;
                border-radius: 25px;
                font-size: 14px;
                outline: none;
                transition: border-color 0.3s ease;
            }

            .input-field:focus {
                border-color: #667eea;
            }

            .send-button {
                padding: 12px 20px;
                background: linear-gradient(135deg, #667eea, #764ba2);
                color: white;
                border: none;
                border-radius: 25px;
                cursor: pointer;
                font-size: 14px;
                font-weight: 600;
                transition: transform 0.2s ease;
            }

            .send-button:hover {
                transform: translateY(-2px);
            }

            .send-button:disabled {
                opacity: 0.5;
                cursor: not-allowed;
                transform: none;
            }

            .loading {
                display: none;
                text-align: center;
                padding: 20px;
                color: #666;
            }

            .loading::after {
                content: '...';
                animation: dots 1.5s steps(4, end) infinite;
            }

            @keyframes dots {
                0%, 20% { color: rgba(0,0,0,0); text-shadow: .25em 0 0 rgba(0,0,0,0), .5em 0 0 rgba(0,0,0,0); }
                40% { color: #666; text-shadow: .25em 0 0 rgba(0,0,0,0), .5em 0 0 rgba(0,0,0,0); }
                60% { text-shadow: .25em 0 0 #666, .5em 0 0 rgba(0,0,0,0); }
                80%, 100% { text-shadow: .25em 0 0 #666, .5em 0 0 #666; }
            }

            .examples {
                margin-top: 15px;
            }

            .examples h3 {
                font-size: 14px;
                color: #666;
                margin-bottom: 10px;
            }

            .example-buttons {
                display: flex;
                flex-wrap: wrap;
                gap: 8px;
            }

            .example-btn {
                padding: 6px 12px;
                background: #f0f0f0;
                border: 1px solid #ddd;
                border-radius: 15px;
                font-size: 12px;
                cursor: pointer;
                transition: all 0.2s ease;
            }

            .example-btn:hover {
                background: #e0e0e0;
                border-color: #bbb;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <div class="logo">
                    <img src="/static/alquimistas.jpg" alt="Logo" onerror="this.style.display='none'">
                </div>
                <h1>Assistente IA Empresarial</h1>
                <p>Consulte dados de funcionários, planilhas e documentos sindicais</p>
            </div>

            <div class="chat-container">
                <div class="messages" id="messages">
                    <div class="message bot-message">
                        Olá! Sou seu assistente para consultas de dados empresariais com processamento de dados REAIS.
                        <br><br>Posso ajudar com informações sobre:
                        <br>• Funcionários ativos, desligados, férias
                        <br>• Dados de sindicatos por estado
                        <br>• Relatórios e planilhas diversas
                        <br>• Consolidação VR com dados reais das bases
                        <br><br>Como posso ajudar você hoje?
                    </div>
                </div>

                <div class="loading" id="loading">
                    Processando sua pergunta com dados reais
                </div>
            </div>

            <div class="input-container">
                <div class="input-group">
                    <input 
                        type="text" 
                        class="input-field" 
                        id="messageInput" 
                        placeholder="Digite sua pergunta..."
                        maxlength="500"
                    >
                    <button class="send-button" id="sendButton" onclick="sendMessage()">
                        Enviar
                    </button>
                </div>

                <div class="examples">
                    <h3>Exemplos de perguntas:</h3>
                    <div class="example-buttons">
                        <button class="example-btn" onclick="setExample('Quantos funcionários estão ativos?')">
                            Funcionários ativos
                        </button>
                        <button class="example-btn" onclick="setExample('Quais são as regras do sindicato de SP?')">
                            Sindicato SP
                        </button>
                        <button class="example-btn" onclick="setExample('Funcionários em férias')">
                            Férias
                        </button>
                        <button class="example-btn" onclick="setExample('Gerar planilha consolidada de VR com dados reais')">
                            Gerar Consolidado VR Real
                        </button>
                    </div>
                </div>
            </div>
        </div>

        <script>
            const messagesContainer = document.getElementById('messages');
            const messageInput = document.getElementById('messageInput');
            const sendButton = document.getElementById('sendButton');
            const loading = document.getElementById('loading');

            messageInput.addEventListener('keypress', function(e) {
                if (e.key === 'Enter' && !e.shiftKey) {
                    e.preventDefault();
                    sendMessage();
                }
            });

            function setExample(text) {
                messageInput.value = text;
                messageInput.focus();
            }

            async function sendMessage() {
                const message = messageInput.value.trim();
                if (!message) return;

                addMessage(message, 'user');
                messageInput.value = '';

                sendButton.disabled = true;
                loading.style.display = 'block';

                try {
                    const response = await fetch('/chat', {
                        method: 'POST',
                        headers: {
                            'Content-Type': 'application/json',
                        },
                        body: JSON.stringify({ message: message })
                    });

                    const data = await response.json();

                    if (data.success) {
                        addMessage(data.response, 'bot');
                    } else {
                        addMessage('Desculpe, ocorreu um erro: ' + data.error, 'bot');
                    }

                } catch (error) {
                    console.error('Erro:', error);
                    addMessage('Erro de conexão. Tente novamente.', 'bot');
                } finally {
                    sendButton.disabled = false;
                    loading.style.display = 'none';
                }
            }

            function addMessage(text, sender) {
                const messageDiv = document.createElement('div');
                messageDiv.className = `message ${sender}-message`;

                const formattedText = text.replace(/\\n/g, '<br>');
                messageDiv.innerHTML = formattedText;

                messagesContainer.appendChild(messageDiv);
                messagesContainer.scrollTop = messagesContainer.scrollHeight;
            }

            window.onload = function() {
                messageInput.focus();
            };
        </script>
    </body>
    </html>
    """

    @app.route("/")
    def home():
        return render_template_string(HTML_TEMPLATE)

    @app.route("/static/<filename>")
    def static_files(filename):
        return send_from_directory(".", filename)

    @app.route("/chat", methods=["POST"])
    def chat():
        try:
            data = request.get_json()
            pergunta = data.get("message", "")

            if not pergunta:
                return jsonify({"success": False, "error": "Pergunta não fornecida"})

            resposta = leitor.processar_pergunta_usuario(pergunta)

            return jsonify({"success": True, "response": resposta})

        except Exception as e:
            return jsonify({"success": False, "error": str(e)})

    import webbrowser
    import threading

    def open_browser():
        webbrowser.open_new("http://localhost:5000")

    print("Iniciando servidor Flask com processamento de dados REAIS...")
    print("Certifique-se de que:")
    print("   - Arquivo 'alquimistas.jpg' está na mesma pasta do script")
    print("   - Pasta 'bases/' contém os arquivos Excel")
    print("   - Pasta 'documents/' contém os arquivos PDF")
    print("   - API Key do Gemini está configurada")
    print("\nAcesse: http://localhost:5000\n")

    threading.Timer(1.5, open_browser).start()
    try:
        # Try running with debug mode first
        app.run(
            debug=True, host="0.0.0.0", port=5000, use_reloader=False, threaded=True
        )
    except ValueError as e:
        if "signal only works in main thread" in str(e):
            print(
                "Debug mode não suportado neste ambiente, executando em modo produção..."
            )
            app.run(
                debug=False,
                host="0.0.0.0",
                port=5000,
                use_reloader=False,
                threaded=True,
            )
        else:
            raise e
